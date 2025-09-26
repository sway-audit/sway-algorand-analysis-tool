"""
Report Export Functionality for Algorand Smart Contract Audit Tool
Enterprise-grade PDF and Excel export with professional formatting
"""

import os
import io
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional
import logging

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Excel Generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.drawing.image import Image as ExcelImage

logger = logging.getLogger(__name__)

class ReportExporter:
    """Professional report export functionality"""
    
    def __init__(self):
        self.export_dir = os.getenv("EXPORT_DIR", "./exports")
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_pdf_report(self, audit_data: Dict[str, Any]) -> str:
        """
        Generate professional PDF audit report
        
        Args:
            audit_data: Complete audit analysis data
            
        Returns:
            Path to generated PDF file
        """
        try:
            filename = f"audit_report_{uuid.uuid4().hex[:8]}.pdf"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.darkblue,
                alignment=1  # Center alignment
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # Title page
            story.append(Paragraph("Smart Contract Security Audit Report", title_style))
            story.append(Spacer(1, 20))
            
            # Executive summary
            story.append(Paragraph("Executive Summary", heading_style))
            summary_data = self._create_executive_summary(audit_data)
            story.append(Paragraph(summary_data, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Contract information
            story.append(Paragraph("Contract Information", heading_style))
            contract_table = self._create_contract_info_table(audit_data)
            story.append(contract_table)
            story.append(Spacer(1, 20))
            
            # Security overview
            story.append(Paragraph("Security Overview", heading_style))
            security_chart = self._create_security_chart(audit_data)
            story.append(security_chart)
            story.append(Spacer(1, 20))
            
            # Findings summary
            story.append(Paragraph("Findings Summary", heading_style))
            findings_table = self._create_findings_summary_table(audit_data)
            story.append(findings_table)
            story.append(Spacer(1, 20))
            
            # Detailed findings
            story.append(Paragraph("Detailed Findings", heading_style))
            detailed_findings = self._create_detailed_findings(audit_data, styles)
            story.extend(detailed_findings)
            
            # Recommendations
            story.append(Paragraph("Recommendations", heading_style))
            recommendations = self._create_recommendations(audit_data, styles)
            story.extend(recommendations)
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"PDF report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"PDF export failed: {str(e)}")
            raise ValueError(f"Failed to generate PDF report: {str(e)}")
    
    def export_excel_report(self, audit_data: Dict[str, Any]) -> str:
        """
        Generate comprehensive Excel audit report
        
        Args:
            audit_data: Complete audit analysis data
            
        Returns:
            Path to generated Excel file
        """
        try:
            filename = f"audit_report_{uuid.uuid4().hex[:8]}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create worksheets
            self._create_summary_sheet(wb, audit_data)
            self._create_findings_sheet(wb, audit_data)
            self._create_metrics_sheet(wb, audit_data)
            self._create_recommendations_sheet(wb, audit_data)
            
            # Save workbook
            wb.save(filepath)
            
            logger.info(f"Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise ValueError(f"Failed to generate Excel report: {str(e)}")
    
    def export_json_report(self, audit_data: Dict[str, Any]) -> str:
        """
        Export structured JSON report for API consumption
        
        Args:
            audit_data: Complete audit analysis data
            
        Returns:
            Path to generated JSON file
        """
        try:
            import json
            
            filename = f"audit_report_{uuid.uuid4().hex[:8]}.json"
            filepath = os.path.join(self.export_dir, filename)
            
            # Prepare structured data
            structured_data = {
                "metadata": {
                    "generated_at": datetime.utcnow().isoformat(),
                    "tool_version": "1.0.0",
                    "export_format": "json"
                },
                "contract": audit_data.get("contract_info", {}),
                "analysis": audit_data.get("analysis_results", {}),
                "findings": audit_data.get("findings", []),
                "metrics": audit_data.get("metrics", {}),
                "recommendations": audit_data.get("recommendations", [])
            }
            
            # Write JSON file
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(structured_data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"JSON report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"JSON export failed: {str(e)}")
            raise ValueError(f"Failed to generate JSON report: {str(e)}")
    
    def _create_executive_summary(self, audit_data: Dict[str, Any]) -> str:
        """Create executive summary text"""
        analysis_report = audit_data.get("analysisReport", {})
        findings = analysis_report.get("findings", [])
        
        total_findings = len(findings)
        risk_score = analysis_report.get("overallRiskScore", "Unknown")
        
        severity_counts = {}
        for finding in findings:
            severity = finding.get("severity", "Unknown")
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        summary = f"""
        This report presents the results of a comprehensive security audit performed on the smart contract.
        The analysis identified {total_findings} security findings with an overall risk assessment of {risk_score}.
        
        Key findings include:
        """
        
        for severity, count in severity_counts.items():
            summary += f"• {count} {severity} severity issues\n"
        
        summary += f"""
        
        The contract was analyzed using advanced static analysis techniques, including vulnerability pattern matching,
        complexity analysis, and gas optimization assessment. All findings have been categorized by severity and
        include detailed remediation guidance.
        """
        
        return summary
    
    def _create_contract_info_table(self, audit_data: Dict[str, Any]) -> Table:
        """Create contract information table"""
        analysis_report = audit_data.get("analysisReport", {})
        
        data = [
            ["Property", "Value"],
            ["Contract Name", analysis_report.get("fileName", "Unknown")],
            ["Analysis Date", analysis_report.get("timestamp", "Unknown")],
            ["Risk Score", analysis_report.get("overallRiskScore", "Unknown")],
            ["Total Findings", str(len(analysis_report.get("findings", [])))]
        ]
        
        table = Table(data, colWidths=[2*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_security_chart(self, audit_data: Dict[str, Any]) -> Drawing:
        """Create security findings pie chart"""
        findings = audit_data.get("analysisReport", {}).get("findings", [])
        
        severity_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0}
        for finding in findings:
            severity = finding.get("severity", "Unknown")
            if severity in severity_counts:
                severity_counts[severity] += 1
        
        # Create pie chart
        drawing = Drawing(400, 200)
        pie = Pie()
        pie.x = 50
        pie.y = 50
        pie.width = 100
        pie.height = 100
        
        pie.data = list(severity_counts.values())
        pie.labels = list(severity_counts.keys())
        pie.slices.strokeWidth = 0.5
        
        # Color scheme
        colors_list = [colors.red, colors.orange, colors.yellow, colors.lightblue, colors.lightgreen]
        for i, color in enumerate(colors_list):
            if i < len(pie.slices):
                pie.slices[i].fillColor = color
        
        drawing.add(pie)
        return drawing
    
    def _create_findings_summary_table(self, audit_data: Dict[str, Any]) -> Table:
        """Create findings summary table"""
        findings = audit_data.get("analysisReport", {}).get("findings", [])
        
        # Count by severity
        severity_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0}
        for finding in findings:
            severity = finding.get("severity", "Unknown")
            if severity in severity_counts:
                severity_counts[severity] += 1
        
        data = [["Severity", "Count", "Percentage"]]
        total = len(findings) if findings else 1
        
        for severity, count in severity_counts.items():
            percentage = f"{(count / total * 100):.1f}%"
            data.append([severity, str(count), percentage])
        
        table = Table(data, colWidths=[1.5*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_detailed_findings(self, audit_data: Dict[str, Any], styles) -> List:
        """Create detailed findings section"""
        findings = audit_data.get("analysisReport", {}).get("findings", [])
        story = []
        
        for i, finding in enumerate(findings, 1):
            # Finding header
            header = f"{i}. {finding.get('vulnerabilityName', 'Unknown Vulnerability')}"
            story.append(Paragraph(header, styles['Heading3']))
            
            # Finding details
            details = f"""
            <b>Severity:</b> {finding.get('severity', 'Unknown')}<br/>
            <b>Line:</b> {finding.get('lineNumber', 'Unknown')}<br/>
            <b>Description:</b> {finding.get('description', 'No description available')}<br/>
            <b>Recommendation:</b> {finding.get('recommendedFix', 'No recommendation available')}
            """
            story.append(Paragraph(details, styles['Normal']))
            story.append(Spacer(1, 12))
        
        return story
    
    def _create_recommendations(self, audit_data: Dict[str, Any], styles) -> List:
        """Create recommendations section"""
        story = []
        
        recommendations = [
            "Review and address all Critical and High severity findings immediately",
            "Implement comprehensive testing including edge cases and attack scenarios",
            "Consider formal verification for critical contract functions",
            "Establish a bug bounty program for ongoing security assessment",
            "Implement monitoring and alerting for contract interactions",
            "Regular security audits should be conducted for contract updates"
        ]
        
        for i, rec in enumerate(recommendations, 1):
            story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
            story.append(Spacer(1, 6))
        
        return story
    
    def _create_summary_sheet(self, wb: Workbook, audit_data: Dict[str, Any]):
        """Create Excel summary sheet"""
        ws = wb.create_sheet("Summary")
        
        # Headers
        ws['A1'] = "Smart Contract Security Audit Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Contract info
        analysis_report = audit_data.get("analysisReport", {})
        ws['A3'] = "Contract Name:"
        ws['B3'] = analysis_report.get("fileName", "Unknown")
        ws['A4'] = "Analysis Date:"
        ws['B4'] = analysis_report.get("timestamp", "Unknown")
        ws['A5'] = "Overall Risk:"
        ws['B5'] = analysis_report.get("overallRiskScore", "Unknown")
        
        # Apply formatting
        for row in range(3, 6):
            ws[f'A{row}'].font = Font(bold=True)
    
    def _create_findings_sheet(self, wb: Workbook, audit_data: Dict[str, Any]):
        """Create Excel findings sheet"""
        ws = wb.create_sheet("Findings")
        
        # Headers
        headers = ["ID", "Vulnerability", "Severity", "Line", "Description", "Recommendation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        findings = audit_data.get("analysisReport", {}).get("findings", [])
        for row, finding in enumerate(findings, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=finding.get("vulnerabilityName", ""))
            ws.cell(row=row, column=3, value=finding.get("severity", ""))
            ws.cell(row=row, column=4, value=finding.get("lineNumber", ""))
            ws.cell(row=row, column=5, value=finding.get("description", ""))
            ws.cell(row=row, column=6, value=finding.get("recommendedFix", ""))
    
    def _create_metrics_sheet(self, wb: Workbook, audit_data: Dict[str, Any]):
        """Create Excel metrics sheet"""
        ws = wb.create_sheet("Metrics")
        
        ws['A1'] = "Security Metrics"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add metrics data here
        metrics = audit_data.get("metrics", {})
        row = 3
        for key, value in metrics.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
    
    def _create_recommendations_sheet(self, wb: Workbook, audit_data: Dict[str, Any]):
        """Create Excel recommendations sheet"""
        ws = wb.create_sheet("Recommendations")
        
        ws['A1'] = "Security Recommendations"
        ws['A1'].font = Font(size=14, bold=True)
        
        recommendations = [
            "Address all Critical and High severity findings",
            "Implement comprehensive testing",
            "Consider formal verification",
            "Establish bug bounty program",
            "Implement monitoring systems",
            "Schedule regular security audits"
        ]
        
        for row, rec in enumerate(recommendations, 3):
            ws.cell(row=row, column=1, value=f"{row-2}.")
            ws.cell(row=row, column=2, value=rec)

# Global exporter instance
report_exporter = ReportExporter()
